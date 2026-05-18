"""电能（功率）约束检查。

实验要求：“在某一时刻电能消耗的上限为 1000W”。
本模块根据任务安排计算任意时刻的总功率，若存在时刻超过 1000W 则判定冲突。
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .data_io import parse_datetime

MAX_POWER = 1000  # W


def _read_event_powers(path: Path | None = None) -> dict[int, int]:
    """从 event.xlsx 读取每个任务 ID 对应的功率（W）。

    默认在 ``data`` 目录下查找包含 ``电能`` 和 ``任务ID`` 表头的 Excel 文件。
    """
    if path is None:
        from .data_io import ROOT_DIR

        candidate = ROOT_DIR / "data" / "event.xlsx"
        if not candidate.exists():
            raise FileNotFoundError(f"Cannot find event.xlsx under {ROOT_DIR / 'data'}")
        path = candidate

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        powers: dict[int, int] = {}
        for row in rows:
            if row is None or row[0] is None:
                continue
            tid = int(row[0])
            power = int(row[2])
            powers[tid] = power
        return powers
    finally:
        workbook.close()


def JudgeWelecClash(
    PlanningEvents: list[list[float]],
    events_path: Path | None = None,
    max_power: int = MAX_POWER,
) -> bool:
    """判断规划事件是否违反功率约束。

    Parameters
    ----------
    PlanningEvents :
        规划事件矩阵，每行至少包含 ``[任务ID, 相对开始时间(秒), 相对结束时间(秒), ...]``。
    events_path :
        任务定义 Excel 文件路径（含功率列）。默认读取 ``data/event.xlsx``。
    max_power :
        功率上限（W），默认 1000W。

    Returns
    -------
    bool
        若任一时刻总功率超过 ``max_power`` 则返回 True，否则返回 False。
    """
    powers = _read_event_powers(events_path)

    # Collect all task intervals with their power
    intervals: list[tuple[float, float, int]] = []
    for event in PlanningEvents:
        tid = int(event[0])
        start = float(event[1])
        end = float(event[2])
        power = powers.get(tid, 0)
        intervals.append((start, end, power))

    if not intervals:
        return False

    # Event-driven sweep: sort all start and end points
    events_list: list[tuple[float, int]] = []
    for start, end, power in intervals:
        events_list.append((start, power))
        events_list.append((end, -power))

    events_list.sort(key=lambda x: x[0])

    current_power = 0
    for _, delta in events_list:
        current_power += delta
        if current_power > max_power:
            return True

    return False
