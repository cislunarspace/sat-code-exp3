"""航天任务规划约束检查的 Python 入口。

本脚本读取规划事件 Excel 表，调用异常区约束和轨道阳光角约束检查函数，
并输出每类约束是否发生冲突。它对应 MATLAB 版本中的主流程，但把数据读取、
异常区判定和阳光角判定拆分到独立函数中，便于测试和替换输入文件。
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.JudgeAvoidAreaClash import JudgeAvoidAreaClash
from src.JudgeSolarClash import JudgeSolarClash

ROOT_DIR = Path(__file__).resolve().parents[1]


def read_planning_events(path: Path | None = None) -> list[list[float]]:
    """读取规划事件表并转换为约束检查所需的数值矩阵。

    默认读取 `data/planningevents.xlsx`，也可以通过 `path` 指定其他 Excel 文件。
    表格第 1 行视为表头，从第 2 行开始读取；空行或首列为空的行会被跳过。
    每条事件只取前 5 列并转为 float，保持与 MATLAB 版本 PlanningEvents 矩阵一致：
    第 1 列为事件编号或类型标识，第 2、3 列为相对开始和结束时间（秒），
    第 4、5 列为阳光角约束的最小和最大绝对值（度）。
    """
    workbook_path = path or ROOT_DIR / "data" / "planningevents.xlsx"
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    sheet = workbook.active
    rows: list[list[float]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row is None or row[0] is None:
            continue
        rows.append([float(value) for value in row[:5]])
    return rows


def main() -> tuple[bool, bool]:
    """执行两类已实现约束检查并返回冲突结果。

    返回值为 `(avoid_area_clash, solar_clash)`：第一个布尔值表示是否存在事件与
    异常区窗口重叠，第二个布尔值表示是否存在事件违反轨道阳光角约束。
    同时将结果打印到标准输出，方便直接运行脚本时查看判定结果。
    """
    planning_events = read_planning_events()
    avoid_area_clash = JudgeAvoidAreaClash(planning_events)
    solar_clash = JudgeSolarClash(planning_events)
    print(f"AvoidAreaClash: {avoid_area_clash}")
    print(f"SolarClash: {solar_clash}")
    return avoid_area_clash, solar_clash


if __name__ == "__main__":
    main()
